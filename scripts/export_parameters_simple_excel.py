"""
Export companies_parameters_only.json to Excel with layout:
Company_Code | file_name | parameters (newline-separated) | count
"""

import json
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment

DATA_DIR = Path("d:/Work/LUMIVST/backend/parameters_analysis")
JSON_PATH = DATA_DIR / "companies_parameters_only.json"
EXCEL_PATH = DATA_DIR / "companies_parameters_simple.xlsx"

with open(JSON_PATH, 'r', encoding='utf-8') as f:
    companies = json.load(f)

rows = []
for code in sorted(companies.keys()):
    entry = companies[code]
    params = entry.get('combined_parameters', entry.get('parameters', []))
    oldest_file = entry.get('oldest_file_name', '')
    newest_file = entry.get('newest_file_name', entry.get('file_name', ''))
    params_text = "\n".join(params)
    rows.append({
        'Company_Code': code,
        'Oldest_File': oldest_file,
        'Newest_File': newest_file,
        'Oldest_Count': len(entry.get('oldest_parameters', [])),
        'Newest_Count': len(entry.get('newest_parameters', [])),
        'Combined_Count': len(params),
        'New_Parameters': ", ".join(entry.get('new_parameters', [])),
        'Parameters': params_text
    })

df = pd.DataFrame(rows, columns=['Company_Code', 'Oldest_File', 'Newest_File', 'Oldest_Count', 'Newest_Count', 'Combined_Count', 'New_Parameters', 'Parameters'])

# Save to Excel
with pd.ExcelWriter(EXCEL_PATH, engine='openpyxl') as writer:
    df.to_excel(writer, sheet_name='Parameters', index=False)

# Enable wrap text for the parameters column
wb = load_workbook(EXCEL_PATH)
ws = wb['Parameters']
# Find parameters column index
for idx, cell in enumerate(ws[1], 1):
    if cell.value == 'parameters':
        param_col = idx
        break
else:
    param_col = 3

for row in ws.iter_rows(min_row=2, min_col=param_col, max_col=param_col):
    for cell in row:
        cell.alignment = Alignment(wrap_text=True, vertical='top')

# Optionally set column widths
ws.column_dimensions['A'].width = 12
ws.column_dimensions['B'].width = 30
ws.column_dimensions[ws.cell(row=1, column=param_col).column_letter].width = 80
ws.column_dimensions['D'].width = 8

wb.save(EXCEL_PATH)
print(f"Created: {EXCEL_PATH}")
