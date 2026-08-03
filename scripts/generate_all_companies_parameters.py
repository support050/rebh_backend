"""
generate_all_companies_parameters.py
=====================================
Extract parameters for each company from its first year (oldest file) and last year (newest file)
and clearly distinguish between:
 1. Parameters in Both Years
 2. Parameters in First Year Only (Removed/Discontinued)
 3. Parameters in Last Year Only (New Parameters Added)
"""

import os
from pathlib import Path
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

DATA_DIR = Path("d:/Work/LUMIVST/backend/data/downloads")
OUTPUT_DIR = Path("d:/Work/LUMIVST/backend/parameters_analysis")
OUTPUT_DIR.mkdir(exist_ok=True)

def extract_parameter_names_from_xls(file_path):
    """Extract actual parameter names/labels from Column 0 of the Excel file"""
    try:
        df = pd.read_excel(file_path, sheet_name=0)
        if df.empty or df.shape[1] == 0:
            return []
        col0 = df.iloc[:, 0].dropna()
        names = []
        seen = set()
        for val in col0:
            s = str(val).replace('\xa0', ' ').replace('\u200b', '').strip()
            if s and s not in ('Start Date', 'End Date', 'nan', 'None', 'Sheet0') and s not in seen:
                seen.add(s)
                names.append(s)
        return names
    except Exception:
        return []

def main():
    print("=" * 80)
    print("Extracting & Distinguishing Parameters (First Year vs Last Year)")
    print("=" * 80)
    
    company_dirs = sorted([d for d in DATA_DIR.iterdir() if d.is_dir()])
    
    summary_rows = []
    long_rows = []
    
    for company_dir in company_dirs:
        company_code = company_dir.name
        
        # Skip duplicate _ar folders for cleaner output
        if company_code.endswith("_ar"):
            continue
            
        xls_files = sorted(list(company_dir.glob("*.xls")) + list(company_dir.glob("*.xlsx")), key=lambda p: p.name)
        
        if not xls_files:
            continue
            
        # Find oldest readable file
        oldest_file = None
        oldest_cols = []
        for f in xls_files:
            cols = extract_parameter_names_from_xls(f)
            if cols:
                oldest_file = f
                oldest_cols = cols
                break
                
        # Find newest readable file
        newest_file = None
        newest_cols = []
        for f in reversed(xls_files):
            cols = extract_parameter_names_from_xls(f)
            if cols:
                newest_file = f
                newest_cols = cols
                break
                
        if not oldest_file or not newest_file:
            continue
            
        set_old = set(oldest_cols)
        set_new = set(newest_cols)
        
        both_params = [p for p in oldest_cols if p in set_new]
        first_only_params = [p for p in oldest_cols if p not in set_new]
        last_only_params = [p for p in newest_cols if p not in set_old]
        
        summary_rows.append({
            "Company_Code": company_code,
            "First_Year_File": oldest_file.name,
            "Last_Year_File": newest_file.name,
            "Both_Years_Count": len(both_params),
            "First_Year_Only_Count": len(first_only_params),
            "Last_Year_Only_Count_New": len(last_only_params),
            "Parameters_Both_Years": "\n".join(both_params),
            "Parameters_First_Year_Only": "\n".join(first_only_params),
            "Parameters_Last_Year_Only_New": "\n".join(last_only_params)
        })
        
        # Long format entries for granular filtering
        for p in both_params:
            long_rows.append({"Company_Code": company_code, "First_Year_File": oldest_file.name, "Last_Year_File": newest_file.name, "Parameter_Name": p, "Status": "Both Years"})
        for p in first_only_params:
            long_rows.append({"Company_Code": company_code, "First_Year_File": oldest_file.name, "Last_Year_File": newest_file.name, "Parameter_Name": p, "Status": "First Year Only"})
        for p in last_only_params:
            long_rows.append({"Company_Code": company_code, "First_Year_File": oldest_file.name, "Last_Year_File": newest_file.name, "Parameter_Name": p, "Status": "Last Year Only (New)"})

        print(f"[{len(summary_rows):3d}] Company {company_code}: Both={len(both_params)} | FirstOnly={len(first_only_params)} | LastOnly(New)={len(last_only_params)}")

    # Clean old files
    for old_file in OUTPUT_DIR.glob("*"):
        try:
            old_file.unlink()
        except Exception:
            pass

    # Save Excel with two tabs (Summary Comparison + Detailed Rows)
    excel_path = OUTPUT_DIR / "company_parameters.xlsx"
    df_summary = pd.DataFrame(summary_rows)
    df_long = pd.DataFrame(long_rows)
    
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        df_summary.to_excel(writer, sheet_name='Parameters_Summary', index=False)
        df_long.to_excel(writer, sheet_name='Parameters_Detailed_List', index=False)
        
    wb = openpyxl.load_workbook(excel_path)
    
    # Format Summary Sheet
    ws = wb['Parameters_Summary']
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    ws.column_dimensions['A'].width = 16  # Company_Code
    ws.column_dimensions['B'].width = 24  # First_Year_File
    ws.column_dimensions['C'].width = 24  # Last_Year_File
    ws.column_dimensions['D'].width = 18  # Both_Years_Count
    ws.column_dimensions['E'].width = 22  # First_Year_Only_Count
    ws.column_dimensions['F'].width = 24  # Last_Year_Only_Count_New
    ws.column_dimensions['G'].width = 75  # Parameters_Both_Years
    ws.column_dimensions['H'].width = 60  # Parameters_First_Year_Only
    ws.column_dimensions['I'].width = 60  # Parameters_Last_Year_Only_New
    
    for row in ws.iter_rows(min_row=2):
        for ci in range(6):
            row[ci].alignment = Alignment(horizontal="center", vertical="top")
        for ci in range(6, 9):
            row[ci].alignment = Alignment(wrap_text=True, vertical="top")
            
    # Format Detailed List Sheet
    ws_long = wb['Parameters_Detailed_List']
    for cell in ws_long[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    ws_long.column_dimensions['A'].width = 16
    ws_long.column_dimensions['B'].width = 24
    ws_long.column_dimensions['C'].width = 24
    ws_long.column_dimensions['D'].width = 80
    ws_long.column_dimensions['E'].width = 24
    
    wb.save(excel_path)
    
    # Save CSV files
    csv_summary_path = OUTPUT_DIR / "company_parameters_summary.csv"
    csv_long_path = OUTPUT_DIR / "company_parameters_detailed.csv"
    
    df_summary.to_csv(csv_summary_path, index=False, encoding='utf-8-sig')
    df_long.to_csv(csv_long_path, index=False, encoding='utf-8-sig')
    
    print("\n" + "=" * 80)
    print(f"SUCCESS: Generated Excel with distinguisher tags at: {excel_path}")
    print(f"Summary CSV: {csv_summary_path}")
    print(f"Detailed CSV: {csv_long_path}")
    print("=" * 80)

if __name__ == "__main__":
    main()
